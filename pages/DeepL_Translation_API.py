import openpyxl
import requests
import datetime as dt
import time
import streamlit as st
from io import BytesIO
import pandas as pd

if "logged_in" not in st.session_state or not st.session_state.logged_in:
    st.error("🚫 You must be logged in to access this page.")
    st.stop()

st.title("DeepL Translation Tool")

with st.sidebar:
    st.markdown("## 👤 Profile")
    st.text_input("User", value=st.session_state.user, key="user_display", disabled=True)

###|-------------------------- PARAMS --------------------------|###
# Only change these when translating


ORIGINAL_LANGUAGE = st.text_input("Type the original language code", help="The code needs to be in all-caps e.g. 'EN'")
TARGET_LANGUAGE = st.text_input("Type the target language code", help="The code needs to be in all-caps e.g. 'EN'")
FILE_TO_TRANSLATE = st.file_uploader("Upload Keyword List: ", type=["xlsx"])
COLUMN_TO_BE_TRANSLATED = 1
COLUMN_TO_TRANSLATE_TO = COLUMN_TO_BE_TRANSLATED + 1
ROW_TO_START_FROM = 2
TIME = dt.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')

###|-------------------------- PARAMS --------------------------|###

# Your DeepL API key (replace with your actual API key)
DEEPL_API_KEY = st.secrets["DEEPL_API_KEY"]

# DeepL API URL for translation
DEEPL_API_URL = 'https://api.deepl.com/v2/translate'  # Use 'api.deepl.com' for paid accounts

# Function to translate text in batches using DeepL API
def translate_batch(texts, target_language=TARGET_LANGUAGE, retries=3):
    attempt = 0
    while attempt < retries:
        try:
            data = {
                'auth_key': DEEPL_API_KEY,
                'text': texts,  # Pass the batch of texts as an array
                'source_lang': ORIGINAL_LANGUAGE,
                'target_lang': target_language
            }
            response = requests.post(DEEPL_API_URL, data=data)
            response.raise_for_status()  # Raise an exception if the request was unsuccessful
            result = response.json()
            translations = [translation['text'] for translation in result['translations']]  # Extract the translations
            return translations  # If successful, return the translations and exit
        except requests.exceptions.RequestException as e:
            print(f"Request error during batch translation (attempt {attempt + 1}): {e}")
            time.sleep(2 ** attempt)  # Exponential backoff
        except KeyError:
            print(f"Unexpected response format during batch translation (attempt {attempt + 1}): {response.text}")
            time.sleep(2 ** attempt)

        attempt += 1  # Increment the attempt counter

    # If all retries fail
    print("Batch translation failed after all retry attempts.")
    return [None] * len(texts)
# Load the Excel file
if FILE_TO_TRANSLATE:
    keywords = pd.read_excel(FILE_TO_TRANSLATE, header=0)
    st.write(keywords.head())
    st.success(f"Loaded {keywords.shape[0]} keywords.")
    file_path = FILE_TO_TRANSLATE  # Replace with the actual file path
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active  # Assuming data is in the first sheet

    # Define the column to translate
    starting_column = COLUMN_TO_TRANSLATE_TO

    # Collect texts from the specified column and track the rows for batch processing
    texts_to_translate = []
    rows_to_update = []
    batch_size = 20  # Reduced batch size to mitigate potential API issues

if st.button("Translate"):
    # Loop through each row in the specified column
    for row in ws.iter_rows(min_row=ROW_TO_START_FROM, min_col=COLUMN_TO_BE_TRANSLATED, max_col=COLUMN_TO_BE_TRANSLATED):
        search_query = row[0].value  # Accessing value in the column
        current_row = row[0].row

        if search_query:
            # Check if the corresponding cell in the target column is empty
            if not ws.cell(row=current_row, column=starting_column).value:
                texts_to_translate.append(search_query)
                rows_to_update.append(current_row)

            # When the batch is full, send it for translation
            if len(texts_to_translate) >= batch_size:
                translations = translate_batch(texts_to_translate, target_language=TARGET_LANGUAGE)

                # Write the translations back into the target column
                for i, translation in enumerate(translations):
                    if translation is not None:
                        ws.cell(row=rows_to_update[i], column=starting_column, value=translation)
                    else:
                        print(f"Translation failed for row {rows_to_update[i]}: {texts_to_translate[i]}")

                # Clear the batch lists
                texts_to_translate.clear()
                rows_to_update.clear()

    # If there are any remaining texts to be translated (less than the batch size)
    if texts_to_translate:
        translations = translate_batch(texts_to_translate, target_language=TARGET_LANGUAGE)
        for i, translation in enumerate(translations):
            if translation is not None:
                ws.cell(row=rows_to_update[i], column=starting_column, value=translation)
            else:
                print(f"Translation failed for row {rows_to_update[i]}: {texts_to_translate[i]}")

    # Save the changes to the Excel file
    filename = f'translated_file_{ORIGINAL_LANGUAGE + "_" + str(TIME)}.xlsx'
    buffer = BytesIO()
    wb.save(buffer)  # Save to a new file or overwrite the original
    st.success("✅ Done! Download your Excel file below:")
    st.download_button("📥 Download Excel", buffer.getvalue(), file_name=filename)